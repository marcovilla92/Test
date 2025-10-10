#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Raybox GUI Enhanced Pro - AMC SYSTEM SRL Edition v2.3
Versione moderna e professionale con font ottimizzati
"""

from __future__ import annotations

import csv
import hashlib
import json
import os
import sys
import threading
import time
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
import tkinter as tk

import requests

# ========== CONFIGURAZIONE AZIENDA ==========
COMPANY_NAME = "AMC SYSTEM SRL"
APP_TITLE = f"Raybox Control Center - {COMPANY_NAME}"
VERSION = "2.3 - Professional Edition"
COMPANY_EMAIL = "info@amcsystem.it"
# ============================================

FONT_FAMILY = "Segoe UI"
FONT_HEADER = (FONT_FAMILY, 28, "bold")
FONT_SUBTITLE = (FONT_FAMILY, 11)
FONT_VERSION = (FONT_FAMILY, 9)
FONT_DATE = (FONT_FAMILY, 8)
FONT_SECTION = (FONT_FAMILY, 10, "bold")
FONT_LABEL = (FONT_FAMILY, 9)
FONT_LABEL_BOLD = (FONT_FAMILY, 9, "bold")
FONT_BUTTON = (FONT_FAMILY, 9, "bold")
FONT_BUTTON_LARGE = (FONT_FAMILY, 11, "bold")
FONT_BUTTON_SMALL = (FONT_FAMILY, 8, "bold")
FONT_STATS = (FONT_FAMILY, 12, "bold")
FONT_MONO = ("Consolas", 10)


def _data_directory() -> Path:
    """Return a platform-appropriate directory for persistent data."""

    if sys.platform.startswith("win"):
        base_dir = Path(os.environ.get("APPDATA", Path.home()))
    else:
        base_dir = Path.home() / ".config"
    data_dir = base_dir / "RayboxControlCenter"
    data_dir.mkdir(parents=True, exist_ok=True)
    return data_dir


DATA_DIR = _data_directory()
TASKS_FILE = DATA_DIR / "raybox_tasks_history.json"


# ------------------ Utility Raybox ------------------

def md5_sign(token: str, secret: str, timestamp_ms: int) -> str:
    """Generate MD5 signature for Raybox API."""

    src = f"timestamp={timestamp_ms}&token={token}&secret={secret}"
    return hashlib.md5(src.encode("utf-8")).hexdigest()


def build_headers(token: str, secret: str) -> dict:
    """Build request headers with authentication."""

    ts = int(time.time() * 1000)
    return {"token": token, "timestamp": str(ts), "sign": md5_sign(token, secret, ts)}


def do_request(
    method: str,
    url: str,
    headers: dict,
    *,
    params=None,
    json_body=None,
    files=None,
    timeout=15,
    dry_run=False,
):
    """Execute HTTP request to Raybox API."""

    if dry_run:
        return {
            "status": "dry-run",
            "msg": "simulazione attiva",
            "url": url,
            "method": method,
        }

    resp = requests.request(
        method,
        url,
        headers=headers,
        params=params,
        json=json_body,
        files=files,
        timeout=timeout,
    )
    resp.raise_for_status()
    try:
        data = resp.json()
    except Exception as exc:  # pragma: no cover - diagnostica
        raise RuntimeError(f"Risposta non-JSON: {resp.text[:500]}") from exc
    if isinstance(data, dict) and data.get("status") not in (0, "dry-run", None):
        raise RuntimeError(f"Errore Raybox: {data.get('msg')}")
    return data


# ------------------ Task Data Model ------------------


class TaskRecord:
    """Rappresenta un task tracciato."""

    def __init__(
        self,
        task_id,
        task_name,
        material="",
        thickness="",
        machine_ip="",
        count=1,
        file_path="",
    ):
        self.task_id = task_id
        self.task_name = task_name
        self.material = material
        self.thickness = thickness
        self.machine_ip = machine_ip
        self.count = count
        self.file_path = file_path
        self.status = "Caricato"
        self.upload_time = datetime.now()
        self.assign_time = None
        self.start_time = None
        self.end_time = None
        self.duration = None
        self.progress = 0
        self.notes = ""

    def to_dict(self):
        return {
            "task_id": self.task_id,
            "task_name": self.task_name,
            "material": self.material,
            "thickness": self.thickness,
            "machine_ip": self.machine_ip,
            "count": self.count,
            "file_path": self.file_path,
            "status": self.status,
            "upload_time": self.upload_time.isoformat() if self.upload_time else None,
            "assign_time": self.assign_time.isoformat() if self.assign_time else None,
            "start_time": self.start_time.isoformat() if self.start_time else None,
            "end_time": self.end_time.isoformat() if self.end_time else None,
            "duration": self.duration,
            "progress": self.progress,
            "notes": self.notes,
        }

    @staticmethod
    def from_dict(data):
        task = TaskRecord(
            data["task_id"],
            data["task_name"],
            data.get("material", ""),
            data.get("thickness", ""),
            data.get("machine_ip", ""),
            data.get("count", 1),
            data.get("file_path", ""),
        )
        task.status = data.get("status", "Caricato")
        task.upload_time = (
            datetime.fromisoformat(data["upload_time"])
            if data.get("upload_time")
            else None
        )
        task.assign_time = (
            datetime.fromisoformat(data["assign_time"])
            if data.get("assign_time")
            else None
        )
        task.start_time = (
            datetime.fromisoformat(data["start_time"])
            if data.get("start_time")
            else None
        )
        task.end_time = (
            datetime.fromisoformat(data["end_time"])
            if data.get("end_time")
            else None
        )
        task.duration = data.get("duration")
        task.progress = data.get("progress", 0)
        task.notes = data.get("notes", "")
        return task

# ------------------ Modern GUI ------------------


class RayboxGUI(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1500x920")
        self.configure(bg="#ecf0f1")
        self.minsize(1300, 750)
        self.option_add("*Font", FONT_LABEL)

        # Connection status
        self.is_connected = False
        self.monitoring_active = False
        self.monitor_thread = None

        # Task tracking
        self.tasks = {}
        self.load_tasks_history()

        # Vars
        self.ip_var = tk.StringVar(value=os.getenv("RAYBOX_IP", "10.1.133.197"))
        self.token_var = tk.StringVar(value=os.getenv("RAYBOX_TOKEN", ""))
        self.secret_var = tk.StringVar(value=os.getenv("RAYBOX_SECRET", ""))
        self.dry_var = tk.BooleanVar(value=True)
        self.app_var = tk.StringVar(value="CypCut")
        self.machine_ip_var = tk.StringVar(value="")
        self.task_id_var = tk.StringVar(value="")
        self.task_name_var = tk.StringVar(value="")
        self.material_var = tk.StringVar(value="")
        self.thick_var = tk.StringVar(value="")
        self.count_var = tk.StringVar(value="1")
        self.target_ip_var = tk.StringVar(value="")
        self.file_path_var = tk.StringVar(value="")

        # Status vars
        self.connection_status_var = tk.StringVar(value="Non connesso")
        self.machine_status_var = tk.StringVar(value="N/A")
        self.task_time_var = tk.StringVar(value="--:--:--")
        self.task_progress_var = tk.StringVar(value="0%")

        self._configure_modern_styles()
        self._build_layout()
        self._refresh_task_table()

    def _configure_modern_styles(self):
        """Configure modern ttk styles."""

        style = ttk.Style()
        available_themes = style.theme_names()
        if "clam" in available_themes:
            style.theme_use("clam")

        style.configure(
            "Modern.TButton",
            font=FONT_BUTTON,
            padding=(18, 8),
            relief="flat",
        )
        style.configure(
            "Action.TButton",
            font=FONT_BUTTON_LARGE,
            padding=(22, 10),
            relief="flat",
        )
        style.configure("Small.TButton", font=FONT_BUTTON_SMALL, padding=(10, 5))

        style.configure("Modern.TLabelframe", background="#ffffff", relief="flat")
        style.configure(
            "Modern.TLabelframe.Label",
            font=FONT_SECTION,
            foreground="#2c3e50",
        )

    def _build_layout(self):
        main_frame = tk.Frame(self, bg="#ecf0f1")
        main_frame.pack(fill="both", expand=True)

        # ========== MODERN HEADER ==========
        header_frame = tk.Frame(main_frame, bg="#1e3a5f", height=85)
        header_frame.pack(fill="x", side="top")
        header_frame.pack_propagate(False)

        left_header = tk.Frame(header_frame, bg="#1e3a5f")
        left_header.pack(side="left", padx=30, pady=15, fill="y")

        tk.Label(
            left_header,
            text=COMPANY_NAME,
            font=FONT_HEADER,
            fg="#ffffff",
            bg="#1e3a5f",
        ).pack(anchor="w")

        tk.Label(
            left_header,
            text="Raybox Control Center",
            font=FONT_SUBTITLE,
            fg="#6ab0f3",
            bg="#1e3a5f",
        ).pack(anchor="w")

        right_header = tk.Frame(header_frame, bg="#1e3a5f")
        right_header.pack(side="right", padx=30, pady=15)

        tk.Label(
            right_header,
            text=f"v{VERSION}",
            font=FONT_VERSION,
            fg="#b4c5d4",
            bg="#1e3a5f",
        ).pack(anchor="e")

        tk.Label(
            right_header,
            text=datetime.now().strftime("%d/%m/%Y"),
            font=FONT_DATE,
            fg="#7f8c8d",
            bg="#1e3a5f",
        ).pack(anchor="e")

        # ========== CONTENT AREA ==========
        content_frame = tk.Frame(main_frame, bg="#ecf0f1")
        content_frame.pack(fill="both", expand=True, padx=15, pady=15)

        self.notebook = ttk.Notebook(content_frame)
        self.notebook.pack(fill="both", expand=True)

        self.operations_tab = tk.Frame(self.notebook, bg="#ecf0f1")
        self.notebook.add(self.operations_tab, text="    ⚙️  Operazioni    ")

        self.tracking_tab = tk.Frame(self.notebook, bg="#ecf0f1")
        self.notebook.add(self.tracking_tab, text="    📊  Tracciamento Task    ")

        self._build_operations_tab()
        self._build_tracking_tab()

        # ========== MODERN FOOTER ==========
        footer_frame = tk.Frame(main_frame, bg="#2c3e50", height=40)
        footer_frame.pack(fill="x", side="bottom")
        footer_frame.pack_propagate(False)

        tk.Label(
            footer_frame,
            text=f"© 2025 {COMPANY_NAME}",
            font=FONT_DATE,
            fg="#ecf0f1",
            bg="#2c3e50",
        ).pack(side="left", padx=20, pady=10)

        tk.Label(
            footer_frame,
            text=f"Supporto: {COMPANY_EMAIL}",
            font=FONT_DATE,
            fg="#5dade2",
            bg="#2c3e50",
        ).pack(side="right", padx=20, pady=10)

    def _build_operations_tab(self):
        main_container = tk.Frame(self.operations_tab, bg="#ecf0f1")
        main_container.pack(fill="both", expand=True, padx=10, pady=10)

        conn_frame = tk.LabelFrame(
            main_container,
            text="  🔌  Connessione Raybox  ",
            font=FONT_SECTION,
            bg="#ffffff",
            fg="#2c3e50",
            relief="solid",
            bd=1,
            padx=20,
            pady=15,
        )
        conn_frame.pack(fill="x", pady=(0, 12))

        inputs_frame = tk.Frame(conn_frame, bg="#ffffff")
        inputs_frame.pack(fill="x", pady=(0, 10))

        row1 = tk.Frame(inputs_frame, bg="#ffffff")
        row1.pack(fill="x", pady=5)

        tk.Label(row1, text="IP Raybox:", font=FONT_LABEL, bg="#ffffff").pack(
            side="left", padx=(0, 8)
        )
        tk.Entry(
            row1,
            textvariable=self.ip_var,
            width=18,
            font=FONT_LABEL,
            relief="solid",
            bd=1,
        ).pack(side="left", padx=5)

        tk.Label(row1, text="TOKEN:", font=FONT_LABEL, bg="#ffffff").pack(
            side="left", padx=(25, 8)
        )
        tk.Entry(
            row1,
            textvariable=self.token_var,
            width=32,
            show="•",
            font=FONT_LABEL,
            relief="solid",
            bd=1,
        ).pack(side="left", padx=5)

        tk.Label(row1, text="SECRET:", font=FONT_LABEL, bg="#ffffff").pack(
            side="left", padx=(25, 8)
        )
        tk.Entry(
            row1,
            textvariable=self.secret_var,
            width=32,
            show="•",
            font=FONT_LABEL,
            relief="solid",
            bd=1,
        ).pack(side="left", padx=5)

        ctrl_row = tk.Frame(conn_frame, bg="#ffffff")
        ctrl_row.pack(fill="x", pady=5)

        self.test_btn = tk.Button(
            ctrl_row,
            text="🔍  Test Connessione",
            command=self.test_connection,
            font=FONT_BUTTON_LARGE,
            bg="#3498db",
            fg="white",
            relief="flat",
            padx=22,
            pady=10,
            cursor="hand2",
            activebackground="#2980b9",
            activeforeground="white",
        )
        self.test_btn.pack(side="left", padx=5)

        tk.Checkbutton(
            ctrl_row,
            text="Dry-run (simula)",
            variable=self.dry_var,
            font=FONT_LABEL,
            bg="#ffffff",
        ).pack(side="left", padx=20)

        tk.Label(
            ctrl_row,
            text="Status:",
            font=FONT_LABEL_BOLD,
            bg="#ffffff",
        ).pack(side="left", padx=(30, 8))
        self.status_label = tk.Label(
            ctrl_row,
            textvariable=self.connection_status_var,
            font=FONT_BUTTON_LARGE,
            fg="#e74c3c",
            bg="#ffffff",
        )
        self.status_label.pack(side="left")

        read_frame = tk.LabelFrame(
            main_container,
            text="  📊  Operazioni di Lettura  ",
            font=FONT_SECTION,
            bg="#ffffff",
            fg="#2c3e50",
            relief="solid",
            bd=1,
            padx=20,
            pady=15,
        )
        read_frame.pack(fill="x", pady=(0, 12))

        read_row = tk.Frame(read_frame, bg="#ffffff")
        read_row.pack(fill="x")

        tk.Button(
            read_row,
            text="📋  Lista Macchine",
            command=self.on_list,
            font=FONT_BUTTON,
            bg="#16a085",
            fg="white",
            relief="flat",
            padx=18,
            pady=7,
            cursor="hand2",
            activebackground="#138d75",
            activeforeground="white",
        ).pack(side="left", padx=5)

        tk.Label(
            read_row,
            text="IP macchina:",
            font=FONT_LABEL,
            bg="#ffffff",
        ).pack(side="left", padx=(25, 8))
        tk.Entry(
            read_row,
            textvariable=self.machine_ip_var,
            width=18,
            font=FONT_LABEL,
            relief="solid",
            bd=1,
        ).pack(side="left", padx=5)

        tk.Label(read_row, text="App:", font=FONT_LABEL, bg="#ffffff").pack(
            side="left", padx=(20, 8)
        )
        ttk.Combobox(
            read_row,
            textvariable=self.app_var,
            values=("CypCut", "TubePro"),
            width=12,
            state="readonly",
            font=FONT_LABEL,
        ).pack(side="left", padx=5)

        tk.Button(
            read_row,
            text="📈  Status Completo",
            command=self.on_status,
            font=FONT_BUTTON,
            bg="#16a085",
            fg="white",
            relief="flat",
            padx=18,
            pady=7,
            cursor="hand2",
            activebackground="#138d75",
            activeforeground="white",
        ).pack(side="left", padx=(20, 5))

        task_frame = tk.LabelFrame(
            main_container,
            text="  ⚙️  Gestione Task  ",
            font=FONT_SECTION,
            bg="#ffffff",
            fg="#2c3e50",
            relief="solid",
            bd=1,
            padx=20,
            pady=15,
        )
        task_frame.pack(fill="x", pady=(0, 12))

        file_row = tk.Frame(task_frame, bg="#ffffff")
        file_row.pack(fill="x", pady=5)

        tk.Label(
            file_row,
            text="File (.zx/.zzx):",
            font=FONT_LABEL,
            bg="#ffffff",
        ).pack(side="left", padx=(0, 8))
        tk.Entry(
            file_row,
            textvariable=self.file_path_var,
            width=65,
            font=FONT_LABEL,
            relief="solid",
            bd=1,
        ).pack(side="left", padx=5, fill="x", expand=True)
        tk.Button(
            file_row,
            text="📁  Sfoglia",
            command=self.pick_file,
            font=FONT_BUTTON,
            bg="#95a5a6",
            fg="white",
            relief="flat",
            padx=16,
            pady=7,
            cursor="hand2",
            activebackground="#7f8c8d",
            activeforeground="white",
        ).pack(side="left", padx=5)

        info_row1 = tk.Frame(task_frame, bg="#ffffff")
        info_row1.pack(fill="x", pady=5)

        tk.Label(info_row1, text="Task ID:", font=FONT_LABEL, bg="#ffffff").pack(
            side="left", padx=(0, 8)
        )
        tk.Entry(
            info_row1,
            textvariable=self.task_id_var,
            width=25,
            font=FONT_LABEL,
            relief="solid",
            bd=1,
        ).pack(side="left", padx=5)

        tk.Label(
            info_row1,
            text="Nome Task:",
            font=FONT_LABEL,
            bg="#ffffff",
        ).pack(side="left", padx=(25, 8))
        tk.Entry(
            info_row1,
            textvariable=self.task_name_var,
            width=30,
            font=FONT_LABEL,
            relief="solid",
            bd=1,
        ).pack(side="left", padx=5)

        info_row2 = tk.Frame(task_frame, bg="#ffffff")
        info_row2.pack(fill="x", pady=5)

        tk.Label(info_row2, text="Materiale:", font=FONT_LABEL, bg="#ffffff").pack(
            side="left", padx=(0, 8)
        )
        tk.Entry(
            info_row2,
            textvariable=self.material_var,
            width=25,
            font=FONT_LABEL,
            relief="solid",
            bd=1,
        ).pack(side="left", padx=5)

        tk.Label(
            info_row2,
            text="Spessore (mm):",
            font=FONT_LABEL,
            bg="#ffffff",
        ).pack(side="left", padx=(25, 8))
        tk.Entry(
            info_row2,
            textvariable=self.thick_var,
            width=12,
            font=FONT_LABEL,
            relief="solid",
            bd=1,
        ).pack(side="left", padx=5)

        tk.Label(info_row2, text="Count:", font=FONT_LABEL, bg="#ffffff").pack(
            side="left", padx=(25, 8)
        )
        tk.Entry(
            info_row2,
            textvariable=self.count_var,
            width=10,
            font=FONT_LABEL,
            relief="solid",
            bd=1,
        ).pack(side="left", padx=5)

        target_row = tk.Frame(task_frame, bg="#ffffff")
        target_row.pack(fill="x", pady=5)

        tk.Label(
            target_row,
            text="IP macchina target:",
            font=FONT_LABEL,
            bg="#ffffff",
        ).pack(side="left", padx=(0, 8))
        tk.Entry(
            target_row,
            textvariable=self.target_ip_var,
            width=25,
            font=FONT_LABEL,
            relief="solid",
            bd=1,
        ).pack(side="left", padx=5)

        btn_frame = tk.Frame(task_frame, bg="#ffffff")
        btn_frame.pack(fill="x", pady=15)

        tk.Button(
            btn_frame,
            text="⬆️  Upload",
            command=self.on_upload,
            font=FONT_BUTTON_LARGE,
            bg="#27ae60",
            fg="white",
            relief="flat",
            padx=28,
            pady=11,
            cursor="hand2",
            activebackground="#229954",
            activeforeground="white",
        ).pack(side="left", padx=8)

        tk.Button(
            btn_frame,
            text="➡️  Assign",
            command=self.on_assign,
            font=FONT_BUTTON_LARGE,
            bg="#2980b9",
            fg="white",
            relief="flat",
            padx=28,
            pady=11,
            cursor="hand2",
            activebackground="#21618c",
            activeforeground="white",
        ).pack(side="left", padx=8)

        tk.Button(
            btn_frame,
            text="❌  Cancel Assign",
            command=self.on_cancel,
            font=FONT_BUTTON_LARGE,
            bg="#e74c3c",
            fg="white",
            relief="flat",
            padx=24,
            pady=11,
            cursor="hand2",
            activebackground="#c0392b",
            activeforeground="white",
        ).pack(side="left", padx=8)

        tk.Button(
            btn_frame,
            text="📂  Open (CypCut)",
            command=self.on_open,
            font=FONT_BUTTON_LARGE,
            bg="#8e44ad",
            fg="white",
            relief="flat",
            padx=24,
            pady=11,
            cursor="hand2",
            activebackground="#7d3c98",
            activeforeground="white",
        ).pack(side="left", padx=8)

        monitor_frame = tk.LabelFrame(
            main_container,
            text="  🔍  Monitoraggio Real-Time  ",
            font=FONT_SECTION,
            bg="#ffffff",
            fg="#2c3e50",
            relief="solid",
            bd=1,
            padx=20,
            pady=15,
        )
        monitor_frame.pack(fill="x", pady=(0, 12))

        ctrl_frame = tk.Frame(monitor_frame, bg="#ffffff")
        ctrl_frame.pack(fill="x", pady=(0, 10))

        self.monitor_btn = tk.Button(
            ctrl_frame,
            text="▶️  Avvia Monitoraggio",
            command=self.toggle_monitoring,
            font=FONT_BUTTON_LARGE,
            bg="#f39c12",
            fg="white",
            relief="flat",
            padx=22,
            pady=10,
            cursor="hand2",
            activebackground="#e67e22",
            activeforeground="white",
        )
        self.monitor_btn.pack(side="left", padx=5)

        tk.Label(
            ctrl_frame,
            text="Status:",
            font=FONT_LABEL_BOLD,
            bg="#ffffff",
        ).pack(side="left", padx=(30, 8))
        self.machine_status_label = tk.Label(
            ctrl_frame,
            textvariable=self.machine_status_var,
            font=FONT_BUTTON_LARGE,
            fg="#7f8c8d",
            bg="#ffffff",
        )
        self.machine_status_label.pack(side="left")

        tk.Label(
            ctrl_frame,
            text="Timer:",
            font=FONT_LABEL_BOLD,
            bg="#ffffff",
        ).pack(side="left", padx=(30, 8))
        self.task_time_label = tk.Label(
            ctrl_frame,
            textvariable=self.task_time_var,
            font=(FONT_FAMILY, 12, "bold"),
            fg="#3498db",
            bg="#ffffff",
        )
        self.task_time_label.pack(side="left")

        tk.Label(
            ctrl_frame,
            text="Progresso:",
            font=FONT_LABEL_BOLD,
            bg="#ffffff",
        ).pack(side="left", padx=(30, 8))
        self.progress_label = tk.Label(
            ctrl_frame,
            textvariable=self.task_progress_var,
            font=FONT_BUTTON_LARGE,
            fg="#27ae60",
            bg="#ffffff",
        )
        self.progress_label.pack(side="left")

        self.progress_bar = ttk.Progressbar(
            monitor_frame, mode="determinate", length=600
        )
        self.progress_bar.pack(fill="x", pady=5)

        log_frame = tk.LabelFrame(
            main_container,
            text="  📝  Log Operazioni  ",
            font=FONT_SECTION,
            bg="#ffffff",
            fg="#2c3e50",
            relief="solid",
            bd=1,
            padx=20,
            pady=15,
        )
        log_frame.pack(fill="both", expand=True)

        log_scroll_frame = tk.Frame(log_frame, bg="#ffffff")
        log_scroll_frame.pack(fill="both", expand=True)

        scrollbar = tk.Scrollbar(log_scroll_frame)
        scrollbar.pack(side="right", fill="y")

        self.log = tk.Text(
            log_scroll_frame,
            height=14,
            wrap="word",
            yscrollcommand=scrollbar.set,
            font=FONT_MONO,
            bg="#f8f9fa",
            relief="flat",
            bd=0,
            padx=10,
            pady=10,
        )
        self.log.pack(side="left", fill="both", expand=True)
        scrollbar.config(command=self.log.yview)

        log_ctrl = tk.Frame(log_frame, bg="#ffffff")
        log_ctrl.pack(fill="x", pady=(10, 0))

        tk.Button(
            log_ctrl,
            text="🗑️  Pulisci",
            command=self.clear_log,
            font=FONT_BUTTON,
            bg="#95a5a6",
            fg="white",
            relief="flat",
            padx=16,
            pady=7,
            cursor="hand2",
            activebackground="#7f8c8d",
            activeforeground="white",
        ).pack(side="left", padx=5)

        tk.Button(
            log_ctrl,
            text="💾  Esporta Log",
            command=self.export_log,
            font=FONT_BUTTON,
            bg="#3498db",
            fg="white",
            relief="flat",
            padx=16,
            pady=7,
            cursor="hand2",
            activebackground="#2980b9",
            activeforeground="white",
        ).pack(side="left", padx=5)

        self._log(f"✅ {COMPANY_NAME} - Raybox Control Center inizializzato")
        self._log(f"ℹ️  Versione {VERSION}")

    def _build_tracking_tab(self):
        main_container = tk.Frame(self.tracking_tab, bg="#ecf0f1")
        main_container.pack(fill="both", expand=True, padx=10, pady=10)

        title_frame = tk.Frame(main_container, bg="#ecf0f1")
        title_frame.pack(fill="x", pady=(0, 12))

        tk.Label(
            title_frame,
            text="📊  Storico Task Processati",
            font=(FONT_FAMILY, 15, "bold"),
            bg="#ecf0f1",
            fg="#2c3e50",
        ).pack(side="left")

        tk.Button(
            title_frame,
            text="🔄  Aggiorna",
            command=self._refresh_task_table,
            font=FONT_BUTTON,
            bg="#3498db",
            fg="white",
            relief="flat",
            padx=16,
            pady=7,
            cursor="hand2",
            activebackground="#2980b9",
            activeforeground="white",
        ).pack(side="right", padx=3)

        tk.Button(
            title_frame,
            text="📊  Excel",
            command=self.export_to_excel,
            font=FONT_BUTTON,
            bg="#27ae60",
            fg="white",
            relief="flat",
            padx=16,
            pady=7,
            cursor="hand2",
            activebackground="#229954",
            activeforeground="white",
        ).pack(side="right", padx=3)

        tk.Button(
            title_frame,
            text="📄  CSV",
            command=self.export_to_csv,
            font=FONT_BUTTON,
            bg="#16a085",
            fg="white",
            relief="flat",
            padx=16,
            pady=7,
            cursor="hand2",
            activebackground="#138d75",
            activeforeground="white",
        ).pack(side="right", padx=3)

        tk.Button(
            title_frame,
            text="🗑️  Pulisci",
            command=self.clear_history,
            font=FONT_BUTTON,
            bg="#e74c3c",
            fg="white",
            relief="flat",
            padx=16,
            pady=7,
            cursor="hand2",
            activebackground="#c0392b",
            activeforeground="white",
        ).pack(side="right", padx=3)

        stats_frame = tk.LabelFrame(
            main_container,
            text="  📈  Statistiche  ",
            font=FONT_SECTION,
            bg="#ffffff",
            fg="#2c3e50",
            relief="solid",
            bd=1,
            padx=24,
            pady=16,
        )
        stats_frame.pack(fill="x", pady=(0, 12))

        stats_grid = tk.Frame(stats_frame, bg="#ffffff")
        stats_grid.pack()

        self.stat_total_label = tk.Label(
            stats_grid,
            text="Totali: 0",
            font=FONT_STATS,
            bg="#ffffff",
            fg="#2c3e50",
        )
        self.stat_total_label.grid(row=0, column=0, padx=30)

        self.stat_completed_label = tk.Label(
            stats_grid,
            text="Completati: 0",
            font=FONT_STATS,
            bg="#ffffff",
            fg="#27ae60",
        )
        self.stat_completed_label.grid(row=0, column=1, padx=30)

        self.stat_processing_label = tk.Label(
            stats_grid,
            text="In Lavorazione: 0",
            font=FONT_STATS,
            bg="#ffffff",
            fg="#3498db",
        )
        self.stat_processing_label.grid(row=0, column=2, padx=30)

        self.stat_avg_time_label = tk.Label(
            stats_grid,
            text="Tempo Medio: --:--:--",
            font=FONT_STATS,
            bg="#ffffff",
            fg="#8e44ad",
        )
        self.stat_avg_time_label.grid(row=0, column=3, padx=30)

        table_frame = tk.LabelFrame(
            main_container,
            text="  📋  Lista Task  ",
            font=FONT_SECTION,
            bg="#ffffff",
            fg="#2c3e50",
            relief="solid",
            bd=1,
            padx=15,
            pady=15,
        )
        table_frame.pack(fill="both", expand=True)

        tree_scroll_y = tk.Scrollbar(table_frame)
        tree_scroll_y.pack(side="right", fill="y")
        tree_scroll_x = tk.Scrollbar(table_frame, orient="horizontal")
        tree_scroll_x.pack(side="bottom", fill="x")

        columns = (
            "task_id",
            "task_name",
            "material",
            "thickness",
            "machine_ip",
            "status",
            "upload_time",
            "start_time",
            "end_time",
            "duration",
            "progress",
        )

        self.task_tree = ttk.Treeview(
            table_frame,
            columns=columns,
            show="headings",
            yscrollcommand=tree_scroll_y.set,
            xscrollcommand=tree_scroll_x.set,
            height=22,
        )

        tree_scroll_y.config(command=self.task_tree.yview)
        tree_scroll_x.config(command=self.task_tree.xview)

        headers_config = {
            "task_id": ("Task ID", 120),
            "task_name": ("Nome Task", 150),
            "material": ("Materiale", 110),
            "thickness": ("Spessore", 90),
            "machine_ip": ("IP Macchina", 125),
            "status": ("Status", 120),
            "upload_time": ("Caricato", 150),
            "start_time": ("Inizio Lavorazione", 150),
            "end_time": ("Fine Lavorazione", 150),
            "duration": ("Durata", 100),
            "progress": ("Progresso %", 95),
        }

        for col, (heading, width) in headers_config.items():
            self.task_tree.heading(col, text=heading)
            self.task_tree.column(
                col,
                width=width,
                anchor="center" if col != "task_name" else "w",
            )

        self.task_tree.pack(fill="both", expand=True)

        self.task_tree.tag_configure("completed", background="#d4edda")
        self.task_tree.tag_configure("processing", background="#cce5ff")
        self.task_tree.tag_configure("assigned", background="#fff3cd")
        self.task_tree.tag_configure("uploaded", background="#f8f9fa")

        self.task_tree.bind("<Button-3>", self._show_task_context_menu)
        self.task_tree.bind("<Double-1>", self._show_task_details)

    # ------------------ Task persistence ------------------

    def load_tasks_history(self):
        if TASKS_FILE.exists():
            try:
                with open(TASKS_FILE, "r", encoding="utf-8") as fh:
                    data = json.load(fh)
                    for task_data in data:
                        task = TaskRecord.from_dict(task_data)
                        self.tasks[task.task_id] = task
            except Exception as exc:  # pragma: no cover - diagnostica
                print(f"Errore caricamento storico: {exc}")

    def save_tasks_history(self):
        try:
            payload = [task.to_dict() for task in self.tasks.values()]
            with open(TASKS_FILE, "w", encoding="utf-8") as fh:
                json.dump(payload, fh, indent=2, ensure_ascii=False)
        except Exception as exc:
            self._log(f"Errore salvataggio storico: {exc}", "ERROR")

    def add_or_update_task(self, task_record):
        self.tasks[task_record.task_id] = task_record
        self.save_tasks_history()
        self._refresh_task_table()

    def _refresh_task_table(self):
        for item in self.task_tree.get_children():
            self.task_tree.delete(item)

        for task in sorted(
            self.tasks.values(),
            key=lambda t: t.upload_time or datetime.min,
            reverse=True,
        ):
            upload_time_str = (
                task.upload_time.strftime("%Y-%m-%d %H:%M:%S")
                if task.upload_time
                else ""
            )
            start_time_str = (
                task.start_time.strftime("%Y-%m-%d %H:%M:%S")
                if task.start_time
                else ""
            )
            end_time_str = (
                task.end_time.strftime("%Y-%m-%d %H:%M:%S")
                if task.end_time
                else ""
            )
            duration_str = task.duration if task.duration else ""

            tag = "uploaded"
            if task.status == "Completato":
                tag = "completed"
            elif task.status == "In lavorazione":
                tag = "processing"
            elif task.status == "Assegnato":
                tag = "assigned"

            self.task_tree.insert(
                "",
                "end",
                values=(
                    task.task_id,
                    task.task_name,
                    task.material,
                    task.thickness,
                    task.machine_ip,
                    task.status,
                    upload_time_str,
                    start_time_str,
                    end_time_str,
                    duration_str,
                    f"{task.progress}%",
                ),
                tags=(tag,),
            )

        self._update_statistics()

    def _update_statistics(self):
        total = len(self.tasks)
        completed = sum(1 for t in self.tasks.values() if t.status == "Completato")
        processing = sum(1 for t in self.tasks.values() if t.status == "In lavorazione")

        completed_tasks = [t for t in self.tasks.values() if t.duration]
        if completed_tasks:
            total_seconds = sum(self._duration_to_seconds(t.duration) for t in completed_tasks)
            avg_seconds = total_seconds / len(completed_tasks)
            avg_time_str = self._format_time(avg_seconds)
        else:
            avg_time_str = "--:--:--"

        self.stat_total_label.config(text=f"Totali: {total}")
        self.stat_completed_label.config(text=f"Completati: {completed}")
        self.stat_processing_label.config(text=f"In Lavorazione: {processing}")
        self.stat_avg_time_label.config(text=f"Tempo Medio: {avg_time_str}")

    def _duration_to_seconds(self, duration_str):
        try:
            hours, minutes, seconds = duration_str.split(":")
            return int(hours) * 3600 + int(minutes) * 60 + int(seconds)
        except Exception:  # pragma: no cover - robustezza
            return 0

    def _show_task_context_menu(self, event):
        item = self.task_tree.identify_row(event.y)
        if item:
            self.task_tree.selection_set(item)
            menu = tk.Menu(self, tearoff=0)
            menu.add_command(label="📋 Dettagli", command=lambda: self._show_task_details(None))
            menu.add_command(label="🗑️ Elimina", command=self._delete_selected_task)
            menu.post(event.x_root, event.y_root)

    def _show_task_details(self, event):
        selection = self.task_tree.selection()
        if not selection:
            return

        item = selection[0]
        values = self.task_tree.item(item)["values"]
        task_id = values[0]

        if task_id in self.tasks:
            task = self.tasks[task_id]
            details = f"""
Task ID: {task.task_id}
Nome: {task.task_name}
Materiale: {task.material}
Spessore: {task.thickness} mm
IP Macchina: {task.machine_ip}
Count: {task.count}
File: {task.file_path}
Status: {task.status}
Progresso: {task.progress}%

Caricato: {task.upload_time.strftime('%Y-%m-%d %H:%M:%S') if task.upload_time else 'N/A'}
Assegnato: {task.assign_time.strftime('%Y-%m-%d %H:%M:%S') if task.assign_time else 'N/A'}
Inizio: {task.start_time.strftime('%Y-%m-%d %H:%M:%S') if task.start_time else 'N/A'}
Fine: {task.end_time.strftime('%Y-%m-%d %H:%M:%S') if task.end_time else 'N/A'}
Durata: {task.duration if task.duration else 'N/A'}

Note: {task.notes if task.notes else 'Nessuna'}

--- {COMPANY_NAME} ---
            """
            messagebox.showinfo("Dettagli Task", details)

    def _delete_selected_task(self):
        selection = self.task_tree.selection()
        if not selection:
            return

        if messagebox.askyesno("Conferma", "Eliminare il task selezionato dallo storico?"):
            item = selection[0]
            values = self.task_tree.item(item)["values"]
            task_id = values[0]

            if task_id in self.tasks:
                del self.tasks[task_id]
                self.save_tasks_history()
                self._refresh_task_table()
                self._log(f"Task '{task_id}' eliminato dallo storico", "INFO")

    def clear_history(self):
        if messagebox.askyesno("Conferma", "Eliminare TUTTO lo storico task?"):
            self.tasks.clear()
            self.save_tasks_history()
            self._refresh_task_table()
            self._log("Storico task pulito", "INFO")

    def export_to_csv(self):
        if not self.tasks:
            messagebox.showwarning("Attenzione", "Nessun task da esportare")
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            initialfile=f"{COMPANY_NAME.replace(' ', '_')}_tasks_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
        )

        if filename:
            try:
                with open(filename, "w", newline="", encoding="utf-8") as fh:
                    writer = csv.writer(fh)
                    writer.writerow(
                        [
                            "Task ID",
                            "Nome Task",
                            "Materiale",
                            "Spessore",
                            "IP Macchina",
                            "Count",
                            "Status",
                            "Caricato",
                            "Assegnato",
                            "Inizio",
                            "Fine",
                            "Durata",
                            "Progresso %",
                            "Note",
                        ]
                    )
                    for task in sorted(
                        self.tasks.values(),
                        key=lambda t: t.upload_time or datetime.min,
                        reverse=True,
                    ):
                        writer.writerow(
                            [
                                task.task_id,
                                task.task_name,
                                task.material,
                                task.thickness,
                                task.machine_ip,
                                task.count,
                                task.status,
                                task.upload_time.strftime("%Y-%m-%d %H:%M:%S")
                                if task.upload_time
                                else "",
                                task.assign_time.strftime("%Y-%m-%d %H:%M:%S")
                                if task.assign_time
                                else "",
                                task.start_time.strftime("%Y-%m-%d %H:%M:%S")
                                if task.start_time
                                else "",
                                task.end_time.strftime("%Y-%m-%d %H:%M:%S")
                                if task.end_time
                                else "",
                                task.duration if task.duration else "",
                                task.progress,
                                task.notes,
                            ]
                        )

                self._log(f"✅ Dati esportati in CSV: {filename}", "SUCCESS")
                messagebox.showinfo("Successo", f"Dati esportati in:\n{filename}")
            except Exception as exc:
                self._log(f"❌ Errore esportazione CSV: {exc}", "ERROR")
                messagebox.showerror("Errore", f"Errore durante l'esportazione:\n{exc}")

    def export_to_excel(self):
        if not self.tasks:
            messagebox.showwarning("Attenzione", "Nessun task da esportare")
            return

        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            messagebox.showerror(
                "Errore",
                "Il modulo 'openpyxl' non è installato.\nInstalla con: pip install openpyxl",
            )
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"{COMPANY_NAME.replace(' ', '_')}_tasks_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        )

        if filename:
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = f"{COMPANY_NAME} Tasks"

                header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")

                headers = [
                    "Task ID",
                    "Nome Task",
                    "Materiale",
                    "Spessore",
                    "IP Macchina",
                    "Count",
                    "Status",
                    "Caricato",
                    "Assegnato",
                    "Inizio",
                    "Fine",
                    "Durata",
                    "Progresso %",
                    "Note",
                ]

                for col, header in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")

                for row, task in enumerate(
                    sorted(
                        self.tasks.values(),
                        key=lambda t: t.upload_time or datetime.min,
                        reverse=True,
                    ),
                    start=2,
                ):
                    ws.cell(row=row, column=1, value=task.task_id)
                    ws.cell(row=row, column=2, value=task.task_name)
                    ws.cell(row=row, column=3, value=task.material)
                    ws.cell(row=row, column=4, value=task.thickness)
                    ws.cell(row=row, column=5, value=task.machine_ip)
                    ws.cell(row=row, column=6, value=task.count)
                    ws.cell(row=row, column=7, value=task.status)
                    ws.cell(
                        row=row,
                        column=8,
                        value=task.upload_time.strftime("%Y-%m-%d %H:%M:%S")
                        if task.upload_time
                        else "",
                    )
                    ws.cell(
                        row=row,
                        column=9,
                        value=task.assign_time.strftime("%Y-%m-%d %H:%M:%S")
                        if task.assign_time
                        else "",
                    )
                    ws.cell(
                        row=row,
                        column=10,
                        value=task.start_time.strftime("%Y-%m-%d %H:%M:%S")
                        if task.start_time
                        else "",
                    )
                    ws.cell(
                        row=row,
                        column=11,
                        value=task.end_time.strftime("%Y-%m-%d %H:%M:%S")
                        if task.end_time
                        else "",
                    )
                    ws.cell(
                        row=row,
                        column=12,
                        value=task.duration if task.duration else "",
                    )
                    ws.cell(row=row, column=13, value=task.progress)
                    ws.cell(row=row, column=14, value=task.notes)

                    if task.status == "Completato":
                        fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif task.status == "In lavorazione":
                        fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    else:
                        fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

                    for col_idx in range(1, 15):
                        ws.cell(row=row, column=col_idx).fill = fill

                for col_cells in ws.columns:
                    max_length = 0
                    column_letter = col_cells[0].column_letter
                    for cell in col_cells:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except Exception:  # pragma: no cover - robustezza
                            pass
                    ws.column_dimensions[column_letter].width = min(max_length + 2, 48)

                wb.save(filename)
                self._log(f"✅ Dati esportati in Excel: {filename}", "SUCCESS")
                messagebox.showinfo(
                    "Successo",
                    f"Dati esportati in:\n{filename}\n\n{COMPANY_NAME}",
                )
            except Exception as exc:
                self._log(f"❌ Errore esportazione Excel: {exc}", "ERROR")
                messagebox.showerror("Errore", f"Errore durante l'esportazione:\n{exc}")

    def export_log(self):
        filename = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
            initialfile=f"{COMPANY_NAME.replace(' ', '_')}_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
        )
        if filename:
            try:
                with open(filename, "w", encoding="utf-8") as fh:
                    fh.write(f"=== {COMPANY_NAME} - Log Operazioni ===\n")
                    fh.write(f"Data export: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                    fh.write(f"Versione: {VERSION}\n")
                    fh.write("=" * 60 + "\n\n")
                    fh.write(self.log.get("1.0", "end"))
                self._log(f"💾 Log esportato: {filename}", "SUCCESS")
                messagebox.showinfo("Successo", f"Log esportato in:\n{filename}")
            except Exception as exc:
                self._log(f"❌ Errore esportazione log: {exc}", "ERROR")
                messagebox.showerror("Errore", f"Errore:\n{exc}")

    def _log(self, msg: str, level="INFO"):
        timestamp = datetime.now().strftime("%H:%M:%S")
        prefix = {"INFO": "ℹ️", "SUCCESS": "✅", "ERROR": "❌", "WARNING": "⚠️"}.get(level, "ℹ️")
        self.log.insert("end", f"[{timestamp}] {prefix}  {msg}\n")
        self.log.see("end")

    def clear_log(self):
        self.log.delete("1.0", "end")
        self._log("🗑️ Log pulito")

    def _get_base(self):
        ip = self.ip_var.get().strip()
        token = self.token_var.get().strip()
        secret = self.secret_var.get().strip()
        if not ip or not token or not secret:
            raise ValueError("Compila IP, TOKEN e SECRET.")
        base = f"http://{ip}:8080"
        return base, token, secret

    def _run_async(self, fn, *args, **kwargs):
        thread = threading.Thread(target=fn, args=args, kwargs=kwargs, daemon=True)
        thread.start()

    def _format_json(self, data):
        return json.dumps(data, indent=2, ensure_ascii=False)

    def test_connection(self):
        def work():
            try:
                self._log("🔍 Test connessione in corso...")
                self.connection_status_var.set("Test in corso...")
                self.status_label.config(fg="#f39c12")

                ip = self.ip_var.get().strip()
                if not ip:
                    raise ValueError("Inserisci l'IP del Raybox")

                base = f"http://{ip}:8080"
                url_time = f"{base}/api/time"
                self._log("Test 1: Verifica raggiungibilità...")

                resp = requests.get(url_time, timeout=5)
                resp.raise_for_status()
                time_data = resp.json()

                if time_data.get("status") == 0:
                    raybox_time = time_data.get("data", "N/A")
                    self._log(f"✅ Raybox raggiungibile! Ora sistema: {raybox_time}", "SUCCESS")
                else:
                    raise RuntimeError("Risposta non valida dal server")

                token = self.token_var.get().strip()
                secret = self.secret_var.get().strip()

                if not token or not secret:
                    self._log("⚠️ TOKEN/SECRET mancanti.", "WARNING")
                    self.connection_status_var.set("⚠️ Parziale")
                    self.status_label.config(fg="#f39c12")
                    self.is_connected = False
                    return

                self._log("Test 2: Verifica autenticazione...")
                headers = build_headers(token, secret)
                url_list = f"{base}/api/datacenter/list"

                resp = requests.get(url_list, headers=headers, timeout=5)
                resp.raise_for_status()
                auth_data = resp.json()

                if auth_data.get("status") == 0:
                    machines = auth_data.get("data", [])
                    self._log(
                        f"✅ Autenticazione riuscita! Trovate {len(machines)} macchine.",
                        "SUCCESS",
                    )
                    self.connection_status_var.set("✅ Connesso")
                    self.status_label.config(fg="#27ae60")
                    self.is_connected = True
                else:
                    raise RuntimeError(f"Errore autenticazione: {auth_data.get('msg')}")

            except Exception as exc:
                self._log(f"❌ Errore: {exc}", "ERROR")
                self.connection_status_var.set("❌ Errore")
                self.status_label.config(fg="#e74c3c")
                self.is_connected = False

        self._run_async(work)

    def toggle_monitoring(self):
        if not self.monitoring_active:
            self.start_monitoring()
        else:
            self.stop_monitoring()

    def start_monitoring(self):
        machine_ip = self.machine_ip_var.get().strip()
        if not machine_ip:
            messagebox.showwarning("Attenzione", "Inserisci l'IP della macchina da monitorare")
            return

        if not self.is_connected:
            messagebox.showwarning("Attenzione", "Effettua prima il test di connessione")
            return

        self.monitoring_active = True
        self.monitor_btn.config(text="⏸️  Ferma Monitoraggio", bg="#e74c3c", activebackground="#c0392b")
        self._log(f"▶️ Monitoraggio avviato per macchina {machine_ip}", "INFO")

        def monitor_loop():
            last_state = None
            current_task_id = None
            task_start_time = None

            while self.monitoring_active:
                try:
                    base, token, secret = self._get_base()
                    headers = build_headers(token, secret)
                    url = f"{base}/api/monitor/cutSystemState"
                    params = {"ip": machine_ip, "appName": self.app_var.get()}

                    if not self.dry_var.get():
                        resp = requests.get(url, headers=headers, params=params, timeout=5)
                        data = resp.json()

                        if data.get("status") == 0:
                            state_data = data.get("data", {}).get("CutSystemState", {})
                            nc_state = state_data.get("NcState", {})

                            sys_state = nc_state.get("SysState", -1)
                            cut_percent = nc_state.get("CutPercent", 0)
                            task_name = nc_state.get("TaskName", "")

                            status_text = self._get_status_text(sys_state)
                            self.machine_status_var.set(status_text)

                            if sys_state == 8:
                                self.machine_status_label.config(fg="#27ae60")
                            elif sys_state == 9:
                                self.machine_status_label.config(fg="#f39c12")
                            else:
                                self.machine_status_label.config(fg="#7f8c8d")

                            if task_name and task_name != current_task_id:
                                current_task_id = task_name
                                if current_task_id in self.tasks:
                                    task = self.tasks[current_task_id]
                                else:
                                    task = TaskRecord(current_task_id, task_name, machine_ip=machine_ip)
                                    self.add_or_update_task(task)

                            if sys_state == 8 and last_state != 8:
                                task_start_time = time.time()
                                if current_task_id and current_task_id in self.tasks:
                                    task = self.tasks[current_task_id]
                                    task.status = "In lavorazione"
                                    task.start_time = datetime.now()
                                    task.machine_ip = machine_ip
                                    self.add_or_update_task(task)
                                self._log(f"⚙️ Task '{task_name}' avviato", "INFO")

                            elif sys_state != 8 and last_state == 8:
                                if task_start_time and current_task_id and current_task_id in self.tasks:
                                    elapsed = time.time() - task_start_time
                                    duration_str = self._format_time(elapsed)
                                    task = self.tasks[current_task_id]
                                    task.status = "Completato"
                                    task.end_time = datetime.now()
                                    task.duration = duration_str
                                    task.progress = 100
                                    self.add_or_update_task(task)
                                    self._log(
                                        f"✅ Task '{current_task_id}' completato in {duration_str}",
                                        "SUCCESS",
                                    )
                                    task_start_time = None

                            if current_task_id and current_task_id in self.tasks and sys_state == 8:
                                task = self.tasks[current_task_id]
                                task.progress = cut_percent
                                self.add_or_update_task(task)

                            self.task_progress_var.set(f"{cut_percent}%")
                            self.progress_bar["value"] = cut_percent

                            if task_start_time and sys_state == 8:
                                elapsed = time.time() - task_start_time
                                self.task_time_var.set(self._format_time(elapsed))
                            elif sys_state != 8:
                                self.task_time_var.set("--:--:--")

                            last_state = sys_state

                    time.sleep(2)

                except Exception as exc:
                    if self.monitoring_active:
                        self._log(f"⚠️ Errore monitoraggio: {exc}", "WARNING")
                    time.sleep(5)

        self.monitor_thread = threading.Thread(target=monitor_loop, daemon=True)
        self.monitor_thread.start()

    def stop_monitoring(self):
        self.monitoring_active = False
        self.monitor_btn.config(text="▶️  Avvia Monitoraggio", bg="#f39c12", activebackground="#e67e22")
        self._log("⏹️ Monitoraggio fermato", "INFO")
        self.machine_status_var.set("N/A")
        self.task_time_var.set("--:--:--")
        self.task_progress_var.set("0%")
        self.progress_bar["value"] = 0

    def _get_status_text(self, sys_state):
        status_map = {
            0: "Standby",
            1: "Simulazione",
            2: "Punto e Movimento",
            3: "Ritorno a zero",
            8: "🟢 In lavorazione",
            9: "⏸️ In pausa",
            10: "Continua",
            -1: "Altro",
        }
        return status_map.get(sys_state, f"Stato {sys_state}")

    def _format_time(self, seconds):
        hours = int(seconds // 3600)
        minutes = int((seconds % 3600) // 60)
        secs = int(seconds % 60)
        return f"{hours:02d}:{minutes:02d}:{secs:02d}"

    def on_list(self):
        def work():
            try:
                base, token, secret = self._get_base()
                headers = build_headers(token, secret)
                url = f"{base}/api/datacenter/list"
                res = do_request("GET", url, headers, dry_run=self.dry_var.get())

                if res.get("status") == 0 or res.get("status") == "dry-run":
                    machines = res.get("data", [])
                    self._log(f"📋 Trovate {len(machines)} macchine:", "SUCCESS")
                    for machine in machines:
                        self._log(f"  • {machine.get('serverName')} - IP: {machine.get('serverIP')}")
                else:
                    self._log(self._format_json(res))
            except Exception as exc:
                self._log(f"Errore: {exc}", "ERROR")

        self._run_async(work)

    def on_status(self):
        def work():
            try:
                base, token, secret = self._get_base()
                machine_ip = self.machine_ip_var.get().strip()
                if not machine_ip:
                    raise ValueError("Inserisci l'IP macchina")
                headers = build_headers(token, secret)
                url = f"{base}/api/monitor/cutSystemState"
                params = {"ip": machine_ip, "appName": self.app_var.get()}
                res = do_request("GET", url, headers, params=params, dry_run=self.dry_var.get())

                self._log("📈 Status macchina:", "SUCCESS")
                self._log(self._format_json(res))
            except Exception as exc:
                self._log(f"Errore: {exc}", "ERROR")

        self._run_async(work)

    def pick_file(self):
        path = filedialog.askopenfilename(
            title="Seleziona file di taglio",
            filetypes=[("Raybox files", "*.zx *.zzx"), ("Tutti i file", "*.*")],
        )
        if path:
            self.file_path_var.set(path)
            self._log(f"📁 File selezionato: {os.path.basename(path)}")

    def on_upload(self):
        def work():
            try:
                base, token, secret = self._get_base()
                file_path = self.file_path_var.get().strip()
                if not file_path or not os.path.isfile(file_path):
                    raise ValueError("Seleziona un file valido (.zx/.zzx)")

                task_id = self.task_id_var.get().strip() or os.path.splitext(os.path.basename(file_path))[0]
                task_name = self.task_name_var.get().strip() or task_id

                headers = build_headers(token, secret)
                url = f"{base}/api/task/upload"

                data = {
                    "taskIdentifier": task_id,
                    "taskName": task_name,
                    "material": self.material_var.get().strip(),
                    "thickness": self.thick_var.get().strip() or "0",
                    "count": self.count_var.get().strip() or "1",
                }

                target_ip = self.target_ip_var.get().strip()
                if target_ip:
                    data["targetMachineIp"] = target_ip

                if not self.dry_var.get():
                    files = {"taskFile": open(file_path, "rb")}
                    try:
                        res = do_request("POST", url, headers, files=files, dry_run=False)
                    finally:
                        files["taskFile"].close()
                else:
                    res = do_request("POST", url, headers, dry_run=True)

                task_record = TaskRecord(
                    task_id=task_id,
                    task_name=task_name,
                    material=self.material_var.get().strip(),
                    thickness=self.thick_var.get().strip(),
                    machine_ip=target_ip,
                    count=int(self.count_var.get().strip() or "1"),
                    file_path=file_path,
                )
                task_record.status = "Caricato"
                self.add_or_update_task(task_record)

                self._log(f"⬆️ Task '{task_name}' uploadato!", "SUCCESS")

                if res.get("status") == 0:
                    returned_id = res.get("data", task_id)
                    self.task_id_var.set(returned_id)

            except Exception as exc:
                self._log(f"Errore: {exc}", "ERROR")

        self._run_async(work)

    def on_assign(self):
        def work():
            try:
                base, token, secret = self._get_base()
                machine_ip = self.machine_ip_var.get().strip()
                task_id = self.task_id_var.get().strip()
                if not machine_ip or not task_id:
                    raise ValueError("Compila IP macchina e Task ID")

                headers = build_headers(token, secret)
                url = f"{base}/api/task/assign"
                body = {"machineIp": machine_ip, "taskIdentifier": task_id}
                do_request("POST", url, headers, json_body=body, dry_run=self.dry_var.get())

                if task_id in self.tasks:
                    task = self.tasks[task_id]
                    task.status = "Assegnato"
                    task.machine_ip = machine_ip
                    task.assign_time = datetime.now()
                    self.add_or_update_task(task)

                self._log(f"➡️ Task '{task_id}' assegnato alla macchina {machine_ip}", "SUCCESS")
            except Exception as exc:
                self._log(f"Errore: {exc}", "ERROR")

        self._run_async(work)

    def on_cancel(self):
        def work():
            try:
                base, token, secret = self._get_base()
                task_id = self.task_id_var.get().strip()
                if not task_id:
                    raise ValueError("Inserisci il Task ID")

                headers = build_headers(token, secret)
                url = f"{base}/api/task/cancelAssign"
                body = {"taskIdentifier": task_id}
                do_request("POST", url, headers, json_body=body, dry_run=self.dry_var.get())

                if task_id in self.tasks:
                    task = self.tasks[task_id]
                    task.status = "Caricato"
                    self.add_or_update_task(task)

                self._log(f"❌ Assegnazione task '{task_id}' cancellata", "SUCCESS")
            except Exception as exc:
                self._log(f"Errore: {exc}", "ERROR")

        self._run_async(work)

    def on_open(self):
        def work():
            try:
                base, token, secret = self._get_base()
                machine_ip = self.machine_ip_var.get().strip()
                task_id = self.task_id_var.get().strip()
                if not machine_ip or not task_id:
                    raise ValueError("Compila IP macchina e Task ID")

                headers = build_headers(token, secret)
                url = f"{base}/api/task/openFile"
                params = {"id": task_id, "ip": machine_ip}
                do_request("GET", url, headers, params=params, dry_run=self.dry_var.get())

                self._log(f"📂 Task '{task_id}' aperto su macchina {machine_ip}", "SUCCESS")
            except Exception as exc:
                self._log(f"Errore: {exc}", "ERROR")

        self._run_async(work)

    def on_closing(self):
        if self.monitoring_active:
            self.stop_monitoring()
        self.destroy()


def main():
    app = RayboxGUI()
    app.protocol("WM_DELETE_WINDOW", app.on_closing)
    app.mainloop()


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        messagebox.showerror(APP_TITLE, f"Errore critico: {exc}")
